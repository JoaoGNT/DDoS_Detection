def resultsL(path):
    import pandas as pd
    import re
    from openpyxl import Workbook

    dataset = []
    wb = Workbook()
    wb.save(filename="C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/roc/roc_svmL.xlsx")
    std = wb['Sheet']
    wb.remove(std)
    with open(path) as f:
        lines = f.readlines()

    for elem in lines:
        if elem =='\n':
            lines.remove(elem)

    for i in range(0,len(lines)-1):
        lines[i] = lines[i].strip()


    for elem in lines:
        if re.match('.*Test:.*', elem) != None or elem == '':
            # title = str(elem.replace(':',''))
            # wb.create_sheet(title=title)
            lines.pop(lines.index(elem))


    lines =list(filter(None, lines))

    #lines.pop()
    splited = []


    for i in range(15):
        start = int(i*len(lines)/15)
        end = int((i+1)*len(lines)/15)
        splited.append(lines[start:end])

    for i in range(0,len(splited)):
        for j in range(0,len(splited[0])):
            splited[i][j] = splited[i][j].partition(':')[2]

    splited_test =[]
    for i in range(5):
        start = int(i*len(splited)/5)
        end = int((i+1)*len(splited)/5)
        splited_test.append(splited[start:end])

    dfvec  = []
    for i in range(0,len(splited_test)):
        # test_vec.append('Test '+format(i))
        test_name = 'Test '+format(i)
        dfvec.append(pd.DataFrame(splited_test[i], columns=['Time(s)', 'C', 'Quant. de Ataques',
                                            'Quant. de Ataques', 'Quant. Ataques Total', 'Acertos', 'Verd-Pos',
                                            'Verd-Neg', 'Erros', 'Falso-Pos', 'Falso-Neg', 'Taxa de Acertos']))
        with pd.ExcelWriter('C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/roc/roc_svmL.xlsx',mode='a') as writer:
            dfvec[i].to_excel(writer, sheet_name=test_name)


def resultsR(path):
    import pandas as pd
    import re
    from openpyxl import Workbook

    dataset = []
    wb = Workbook()
    wb.save(filename="C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/roc/roc_svmR.xlsx")
    std = wb['Sheet']
    wb.remove(std)
    with open(path) as f:
        lines = f.readlines()

    for elem in lines:
        if elem == '\n':
            lines.remove(elem)

    for i in range(0, len(lines) - 1):
        lines[i] = lines[i].strip()

    data_v = []

    for elem in lines:
        if re.match('.*Test:.*', elem) != None or elem == '':
            # title = str(elem.replace(':',''))
            # wb.create_sheet(title=title)
            lines.pop(lines.index(elem))

    lines = list(filter(None, lines))

    lines.pop()
    splited = []

    for i in range(45):
        start = int(i * len(lines)/45)
        end = int((i + 1) * len(lines)/45)
        splited.append(lines[start:end])

    for i in range(0, len(splited)):
        for j in range(0, len(splited[0])):
            splited[i][j] = splited[i][j].partition(':')[2]

    splited_test = []
    for i in range(5):
        start = int(i * len(splited) / 5)
        end = int((i + 1) * len(splited) / 5)
        splited_test.append(splited[start:end])

    dfvec = []
    test_vec = []
    for i in range(0, len(splited_test)):
        # test_vec.append('Test '+format(i))
        test_name = 'Test ' + format(i)
        dfvec.append(pd.DataFrame(splited_test[i], columns=['Time(s)', 'C', 'Gamma', 'Quant. de Ataques',
                                                            'Quant. de Ataques', 'Quant. Ataques Total', 'Acertos',
                                                            'Verd-Pos',
                                                            'Verd-Neg', 'Erros', 'Falso-Pos', 'Falso-Neg',
                                                            'Taxa de Acertos']))
        with pd.ExcelWriter('C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/roc/roc_svmR.xlsx',
                            mode='a') as writer:
            dfvec[i].to_excel(writer, sheet_name=test_name)

def resultsP(path):
    import pandas as pd
    import re
    from openpyxl import Workbook

    dataset = []
    wb = Workbook()
    wb.save(filename="C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/roc/roc_svmP.xlsx")
    std = wb['Sheet']
    wb.remove(std)
    with open(path) as f:
        lines = f.readlines()

    for elem in lines:
        if elem =='\n':
            lines.remove(elem)

    for i in range(0,len(lines)-1):
        lines[i] = lines[i].strip()

    data_v = []

    for elem in lines:
        if re.match('.*Test:.*', elem) != None or elem == '':
            # title = str(elem.replace(':',''))
            # wb.create_sheet(title=title)
            lines.pop(lines.index(elem))


    lines =list(filter(None, lines))

    lines.pop()
    splited = []


    for i in range(45):
        start = int(i*len(lines)/45)
        end = int((i+1)*len(lines)/45)
        splited.append(lines[start:end])

    for i in range(0,len(splited)):
        for j in range(0,len(splited[0])):
            splited[i][j] = splited[i][j].partition(':')[2]

    splited_test =[]
    for i in range(5):
        start = int(i*len(splited)/5)
        end = int((i+1)*len(splited)/5)
        splited_test.append(splited[start:end])

    dfvec  = []
    test_vec=[]
    for i in range(0,len(splited_test)):
        # test_vec.append('Test '+format(i))
        test_name = 'Test '+format(i)
        dfvec.append(pd.DataFrame(splited_test[i], columns=['Time(s)', 'C','Degree', 'Quant. de Ataques',
                                            'Quant. de Ataques', 'Quant. Ataques Total', 'Acertos', 'Verd-Pos',
                                            'Verd-Neg', 'Erros', 'Falso-Pos', 'Falso-Neg', 'Taxa de Acertos']))
        with pd.ExcelWriter('C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/roc/roc_svmP.xlsx',mode='a') as writer:
            dfvec[i].to_excel(writer, sheet_name=test_name)

resultsL('C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/results/SVMresultsL.txt')
resultsR('C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/results/SVMresultsR.txt')
resultsP('C:/Users/jg_te/PycharmProjects/DDoS_Detection/2017/ML/results/SVMresultsP.txt')